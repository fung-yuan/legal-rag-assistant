"""
03_text_cleaner.py
==================
Loads the master page-level dataset (moroccan_law_pages.csv), applies text
normalization, strips header/footer noise, removes decorative dividers, 
and saves:
  1. Cleaned text files to cleaned_text/{year}/{bulletin}/page{N}.txt
  2. A new cleaned master dataset data/moroccan_law_pages_cleaned.csv
  3. A styled Excel sheet data/moroccan_law_pages_cleaned.xlsx
"""

import os
import re
import sys
import logging
import unicodedata
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Config
INPUT_CSV = "data/etl_intermediate/moroccan_law_pages.csv"
OUTPUT_DIR = "data/cleaned_text"
OUTPUT_CSV = "data/etl_intermediate/moroccan_law_pages_cleaned.csv"
OUTPUT_EXCEL = "data/etl_intermediate/moroccan_law_pages_cleaned.xlsx"
LOG_FILE = "data/logs/03_text_cleaner.log"

# Logging setup
os.makedirs("data/logs", exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# Stopword sets for table detection
COMMON_WORDS_FR = {
    'de', 'la', 'le', 'et', 'les', 'en', 'des', 'du', 'un', 'une', 'pour', 'dans', 
    'par', 'sur', 'au', 'aux', 'loi', 'article', 'bulletin', 'officiel', 'est', 'sont'
}

COMMON_WORDS_AR = {
    'من', 'في', 'على', 'إلى', 'الجريدة', 'الرسمية', 'ظهير', 'قانون', 'مرسوم', 'المادة', 'رقم'
}

def is_pure_data_table(text: str, lang: str) -> bool:
    """Detects if a page is a pure scientific/mathematical data table rather than legal text."""
    if not isinstance(text, str) or not text.strip() or len(text.strip()) < 100:
        return False
        
    text_len = len(text)
    
    # 1. Symbol & Digit density (exclude letters and spaces)
    symbols_and_digits = re.findall(r'[^a-zA-Z\s\u00C0-\u017F\u0600-\u06FF]', text)
    density = len(symbols_and_digits) / text_len
    
    # 2. Stopword vocabulary match ratio
    words = re.findall(r'\b\w+\b', text.lower())
    total_words = len(words)
    if total_words == 0:
        return True
        
    common_set = COMMON_WORDS_FR if lang == "FR" else COMMON_WORDS_AR
    matched_words = sum(1 for w in words if w in common_set)
    vocab_match_ratio = matched_words / total_words
    
    # If the stopword frequency is under 5% and digit/symbol density is over 25%, it is a table
    if vocab_match_ratio < 0.05 and density > 0.25:
        return True
    return False

def strip_headers_footers(text: str, lang: str) -> str:
    """Removes repetitive headers, footers, page numbers, and dates from the page text."""
    if not isinstance(text, str):
        return ""
        
    lines = text.splitlines()
    cleaned_lines = []
    
    for line in lines:
        stripped = line.strip()
        
        # Skip empty lines in header/footer checking
        if not stripped:
            cleaned_lines.append(line)
            continue
            
        # ─── French Edition Headers/Footers ───
        if lang == "FR":
            # 1. Date & Number header (e.g. "N° 5514 — 16 rabii I 1428 (5-4-2007)")
            if re.match(r'(?i)^n°\s+\d+\s+—\s+', stripped):
                continue
            # 2. BULLETIN OFFICIEL header (isolated or with page number)
            if re.match(r'(?i)^(bulletin\s+officiel\s+\d+|\d+\s+bulletin\s+officiel|bulletin\s+officiel)$', stripped):
                continue
            # 3. Imprimerie Officielle printer labels
            if re.match(r'(?i)^(imprimerie\s+officielle|rabat\s+-\s+chellah)$', stripped):
                continue
            # 4. Printer numbers/codes
            if re.match(r'^Tél\s*:\s*\d+', stripped) or re.match(r'^Compte\s+n°\s*:', stripped):
                continue
            # 5. Isolated page numbers at the very top or bottom of a line
            if re.match(r'^\d+$', stripped):
                continue
                
        # ─── Arabic Edition Headers/Footers ───
        else:
            # 1. Date & Number header (e.g. "عدد 5514 - 16 ربيع الأول...")
            if re.match(r'^عدد\s+\d+\s+-\s+', stripped):
                continue
            # 2. الجريدة الرسمية header
            if re.match(r'^(الجريدة\s+الرسمية\s+\d+|\d+\s+الجريدة\s+الرسمية|الجريدة\s+الرسمية)$', stripped):
                continue
            # 3. المطبعة الرسمية labels
            if re.match(r'^(المطبعة\s+الرسمية|الرباط\s+-\s+شالة)$', stripped):
                continue
            # 4. Isolated page numbers
            if re.match(r'^\d+$', stripped):
                continue
                
        cleaned_lines.append(line)
        
    return "\n".join(cleaned_lines).strip()

def clean_text_content(text: str, lang: str) -> str:
    """Applies a pipeline of regex cleaning rules to normalize layout and character anomalies."""
    if not isinstance(text, str):
        return ""
        
    # 1. Unicode Normalization (NFC Form standardizes accents and letters)
    cleaned = unicodedata.normalize('NFC', text)
    
    # 2. Remove Signature Block / Layout Dividers (3+ repeats of _, -, ., |)
    cleaned = re.sub(r'_{3,}', ' ', cleaned)
    cleaned = re.sub(r'\-{3,}', ' ', cleaned)
    cleaned = re.sub(r'\.{4,}', ' ', cleaned)  # Dot leaders in index pages
    cleaned = re.sub(r'\|{3,}', ' ', cleaned)
    
    # 3. Standardize Quotation Marks & Apostrophes
    cleaned = re.sub(r'[\u201C\u201D\u201E\u201F\u2033\u2036]', '"', cleaned)
    cleaned = re.sub(r'[\u2018\u2019\u201A\u201B\u2032\u2035]', "'", cleaned)
    
    # 4. De-hyphenation: rejoin words broken across lines
    # e.g., 'com-\nmerce' -> 'commerce'
    cleaned = re.sub(r'(\w+)-\s*\n\s*(\w+)', r'\1\2', cleaned)
    
    # 5. Remove repetitive Header/Footer Lines
    cleaned = strip_headers_footers(cleaned, lang)
    
    # 6. Collapse runs of 3+ blank lines to a single blank line
    cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
    
    return cleaned.strip()

def style_excel(writer, df):
    """Styles the cleaned Excel sheets nicely for review."""
    sheet_name = 'Cleaned Page Text'
    worksheet = writer.sheets[sheet_name]
    worksheet.auto_filter.ref = worksheet.dimensions
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    text_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    for col_idx, col in enumerate(worksheet.columns, 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]
        
        max_len = len(col_name)
        is_content_col = (col_name == "Clean_Content" or col_name == "File_Path")
        
        for cell in col:
            if cell.row > 1:
                cell.font = text_font
                cell.border = thin_border
                
                if col_name in ["Year", "Page", "Language", "Clean_Length", "Word_Count"]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top")
                    
            val_str = str(cell.value or '')
            if not is_content_col:
                max_len = max(max_len, len(val_str))
            else:
                max_len = max(max_len, min(len(val_str), 50))
                
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

def main():
    log.info("=" * 65)
    log.info("Moroccan Law AI — Text Corpus Normalizer and Cleaner")
    log.info("=" * 65)
    
    if not os.path.exists(INPUT_CSV):
        log.error(f"Input file '{INPUT_CSV}' does not exist! Please run 02_generate_excel.py first.")
        sys.exit(1)
        
    log.info(f"Loading raw master dataset '{INPUT_CSV}'...")
    df = pd.read_csv(INPUT_CSV, encoding="utf-8")
    total_rows = len(df)
    log.info(f"Loaded {total_rows} pages to clean.")
    
    cleaned_data = []
    
    log.info("Running text cleaning pipeline on all pages...")
    for idx, row in df.iterrows():
        if idx > 0 and idx % 5000 == 0:
            log.info(f"Cleaned {idx}/{total_rows} pages...")
            
        raw_text = row.get("Content", "")
        lang = row.get("Language", "FR")
        year = str(row.get("Year", "Unknown"))
        bulletin = row.get("Bulletin", "Unknown")
        page_num = row.get("Page", 0)
        
        # Check if this page is a scientific data table and should be filtered out
        if is_pure_data_table(raw_text, lang):
            log.info(f"  [Filtered Table] Skipping scientific/data table: {bulletin} Page {page_num}")
            continue
            
        # Apply cleaner pipeline
        cleaned_text = clean_text_content(raw_text, lang)
        
        # Save the physical cleaned file
        cleaned_file_dir = Path(OUTPUT_DIR) / year / bulletin
        cleaned_file_dir.mkdir(parents=True, exist_ok=True)
        cleaned_file_path = cleaned_file_dir / f"page{page_num}.txt"
        cleaned_file_path.write_text(cleaned_text, encoding="utf-8")
        
        cleaned_data.append({
            "Year": row.get("Year"),
            "Bulletin": bulletin,
            "Page": page_num,
            "Language": lang,
            "Clean_Length": len(cleaned_text),
            "Word_Count": len(cleaned_text.split()),
            "Clean_Content": cleaned_text[:32700], # Excel cell size limit
            "File_Path": str(cleaned_file_path.relative_to(Path(".")))
        })
        
    cleaned_df = pd.DataFrame(cleaned_data)
    cleaned_df = cleaned_df.sort_values(by=["Year", "Bulletin", "Page"]).reset_index(drop=True)
    
    # Save CSV
    log.info(f"Saving cleaned CSV to '{OUTPUT_CSV}'...")
    cleaned_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    
    # Save styled Excel
    log.info(f"Saving styled cleaned Excel to '{OUTPUT_EXCEL}'...")
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        cleaned_df.to_excel(writer, sheet_name='Cleaned Page Text', index=False)
        style_excel(writer, cleaned_df)
        
    log.info("=" * 65)
    log.info("Text cleaning pipeline complete!")
    log.info(f"  Cleaned text files saved in: {os.path.abspath(OUTPUT_DIR)}")
    log.info(f"  Cleaned CSV saved to       : {os.path.abspath(OUTPUT_CSV)}")
    log.info(f"  Cleaned Excel saved to     : {os.path.abspath(OUTPUT_EXCEL)}")
    log.info("=" * 65)

if __name__ == "__main__":
    main()

"""
02_generate_excel.py
====================
Generates a structured, clean, page-level Excel sheet and CSV file
containing the OCR'd and native text from raw_text/.

Columns:
  - Year (int)
  - Bulletin (str)
  - Page (int)
  - Language (str: FR/AR)
  - Text_Length (int)
  - Word_Count (int)
  - Line_Count (int)
  - Content (str: text up to Excel cell limit)
  - File_Path (str: relative path)

Usage:
    python pipeline/etl/02_generate_excel.py
"""

import os
import re
import sys
import logging
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Config
INPUT_DIR = "data/raw_text"
OUTPUT_EXCEL = "data/etl_intermediate/moroccan_law_pages.xlsx"
OUTPUT_CSV = "data/etl_intermediate/moroccan_law_pages.csv"
LOG_FILE = "data/logs/02_generate_excel.log"

# Logging setup
os.makedirs("data/logs", exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

def parse_text_files():
    data = []
    input_path = Path(INPUT_DIR)
    
    if not input_path.exists():
        log.error(f"Input directory '{INPUT_DIR}' does not exist!")
        sys.exit(1)
        
    log.info(f"Scanning '{INPUT_DIR}' for text files...")
    
    text_files = sorted(input_path.rglob("*.txt"))
    total_files = len(text_files)
    log.info(f"Found {total_files} text files to process.")
    
    for i, file_path in enumerate(text_files, 1):
        if i % 1000 == 0 or i == total_files:
            log.info(f"Processed {i}/{total_files} files...")
            
        try:
            rel_parts = file_path.relative_to(input_path).parts
            if len(rel_parts) >= 3:
                year = rel_parts[0]
                bulletin = rel_parts[1]
                page_file = rel_parts[2]
            else:
                year = file_path.parent.parent.name
                bulletin = file_path.parent.name
                page_file = file_path.name
                
            page_match = re.search(r'page(\d+)', page_file)
            page_num = int(page_match.group(1)) if page_match else 0
            
            lang = "AR" if "_Ar" in bulletin else "FR"
            
            text = file_path.read_text(encoding="utf-8").strip()
            
            text_len = len(text)
            word_count = len(text.split())
            line_count = len(text.splitlines())
            
            # Clean illegal XML characters that make openpyxl crash
            cleaned_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
            excel_content = cleaned_text[:32700] if len(cleaned_text) > 32700 else cleaned_text
            
            data.append({
                "Year": int(year) if year.isdigit() else 0,
                "Bulletin": bulletin,
                "Page": page_num,
                "Language": lang,
                "Text_Length": text_len,
                "Word_Count": word_count,
                "Line_Count": line_count,
                "Content": excel_content,
                "File_Path": str(file_path.relative_to(Path(".")))
            })
        except Exception as e:
            log.error(f"Error reading file '{file_path}': {e}")
            
    return pd.DataFrame(data)

def style_excel(writer, df):
    workbook = writer.book
    sheet_name = 'Page Level Text'
    worksheet = writer.sheets[sheet_name]
    
    worksheet.auto_filter.ref = worksheet.dimensions
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    text_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    for col_idx, col in enumerate(worksheet.columns, 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]
        
        max_len = len(col_name)
        is_content_col = (col_name == "Content" or col_name == "File_Path")
        
        for cell in col:
            if cell.row > 1:
                cell.font = text_font
                cell.border = thin_border
                
                if col_name in ["Year", "Page", "Language", "Text_Length", "Word_Count", "Line_Count"]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top")
                    
            val_str = str(cell.value or '')
            if not is_content_col:
                max_len = max(max_len, len(val_str))
            else:
                max_len = max(max_len, min(len(val_str), 50))
                
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

def main():
    log.info("=" * 65)
    log.info("Moroccan Law AI — Page-Level Excel and CSV Dataset Generator")
    log.info("=" * 65)
    
    df = parse_text_files()
    
    if df.empty:
        log.warning("No text content found to export!")
        sys.exit(0)
        
    df = df.sort_values(by=["Year", "Bulletin", "Page"]).reset_index(drop=True)
    
    log.info(f"Extracted data for {len(df)} pages. Saving files...")
    
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    log.info(f"CSV saved to: {os.path.abspath(OUTPUT_CSV)}")
    
    log.info("Saving and styling Excel workbook...")
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Page Level Text', index=False)
        style_excel(writer, df)
        
    log.info(f"Excel workbook saved to: {os.path.abspath(OUTPUT_EXCEL)}")
    log.info("=" * 65)
    log.info("Excel generation complete!")
    log.info("=" * 65)

if __name__ == "__main__":
    main()
